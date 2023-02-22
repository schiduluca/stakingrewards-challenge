import csv
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import re

def write_file(ws):
   with open('transactions_result.csv', 'w') as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])


def read_file():
    wb = openpyxl.Workbook()
    ws = wb.active

    with open('transactions.csv') as f:
        reader = csv.reader(f, delimiter='|')
        for row in reader:
            ws.append(row)

    return ws

def concat(args):
    result = ""
    for arg in args:
        result += arg
    return result
    
# get_cell_reference (A..Z)n operation
def get_cell_reference(text):
    regex = "[A-Z][0-9]"
    return re.findall(regex, str(text))

# get_evaluated_cell A^ operation
def get_evaluated_cell(text):
    regex = "[A-Z]\^[^v]"
    return re.findall(regex, str(text))

# get_delimiter
def get_delimiter(text):
    return re.findall(r'"(.*?)"', str(text))

# get_delimiter
def get_referenced_column(text):
    regex = "@[\w]+\<[\w\s]+\>"
    return re.findall(regex, str(text))

# get_evaluated_cell A^v operation
def get_evaluated_cell_last(text):
    regex = "[A-Z]\^[v]"
    return re.findall(regex, str(text))

def read_excel():
    wb = load_workbook('ResultExcelFile.xlsx', data_only=False)
    ws = wb.active
    return ws


def prepare_formula(text, ws, r, c):
    new = text[1:].replace(" ", "").strip()
    ws.cell(r, c).value = new
    # executes the referencing operation (A..Z)n
    references = get_cell_reference(new)
    for rf in references:
        new = new.replace(rf, ws.cell(row = int(rf[1:]), column = ord(rf[0]) - 64).value)
        ws.cell(r, c).value = new

    
    # executes the evaluated cell operation A^
    evaluated_values = get_evaluated_cell(new)
    for ev in evaluated_values:
        ev = ev[:len(ev) - 1]
        new = new.replace(ev, ws.cell(row = r - 1, column = c).value)
        ws.cell(r, c).value = new

    # executes the get_referenced_column
    references = get_referenced_column(new)
    for ref in references:
        new_ref = ref[1:]
        args = new_ref.split("<")
        column = args[0]
        row = args[1].split(">")[0]
        for ro in ws:
            for cell in ro:
                if cell.value is not None and '!'+column == cell.value:
                    target_row = cell.row + int(row)
                    new = new.replace(ref, ws.cell(target_row, cell.column).value)
                    ws.cell(r, c).value = new

    # executes the operation A^v
    evaluated_values_last = get_evaluated_cell_last(new)
    for evl in evaluated_values_last:
        count = 0
        tempR = r
        while tempR > 1:
            if count > 1 and ws.cell(row = tempR, column = ord(evl[0]) - 64) is not None:
                new = new.replace(evl, ws.cell(row = tempR, column = ord(evl[0]) - 64).value)
                ws.cell(r, c).value = new
                break
            if ws.cell(row = tempR, column = 1).value.startswith("!"):
                count += 1
            tempR -= 1
            
operations = ["concat", "sum", "spread", "split", "text", "incFrom", "bte"]

def evaluate(text):
    while True:
        operation = ""
        latest = -1
        for op in operations:
            x = text.rfind(op)
            if x > latest:
                latest = x
                operation = op      
        if latest == -1:
            if '+' in text:
                return eval(text)
            break
            
        index = text.rindex(operation)
        dataIndex = index + len(operation) + 1
        data_end = text.find(")", index)
        data = text[dataIndex: data_end]
        if operation == "split":
            delimiter = get_delimiter(data)[0]
            arguments = data.split(",\"")[0]
            arguments = arguments.split(delimiter)
            arguments = ','.join(arguments)
            text = text[:index] + arguments + text[data_end+1:]

        elif operation == "spread":
            text = text[:index] + arguments + text[data_end+1:]

        elif operation == "sum":
            arguments = data.split(",")
            sum = 0
            for ar in arguments:
                sum += float(ar)
            text = text[:index] + str(sum) + text[data_end+1:]
        elif operation == "incFrom":
            n = int(data)
            n += 1
            text = text[:index] + str(n) + text[data_end+1:]
        elif operation == "text":
            text = text[:index] + "\"" + data + "\""+ text[data_end+1:]
        elif operation == "concat":
            arguments = data.split(",")
            for i, ar in enumerate(arguments):
                arguments[i] = ar.replace("\"", "")

            result = ''.join(arguments)
            text = text[:index] + result + text[data_end+1:]
        elif operation == "bte":
            args = data.split(",")
            result = max([float(x) for x in args])
            text = text[:index] + str(result) + text[data_end+1:]
    return text

def proccess_spreadsheet(ws):
    # replace ^^ with appropriate formulas
    for row in ws:
        for cell in row:
            if cell.value is not None and '^^' in cell.value:
                if cell.row - 1 >= 0:
                    ws.cell(cell.row, cell.column).value = ws.cell(cell.row - 1, cell.column).value

    for row in ws:
        for cell in row:
            if cell.value is not None and '=' in cell.value:
                prepare_formula(cell.value, ws, cell.row, cell.column)
                result = evaluate(cell.value)
                ws.cell(cell.row, cell.column).value = str(result)
    
ws = read_file()

proccess_spreadsheet(ws)
write_file(ws)
